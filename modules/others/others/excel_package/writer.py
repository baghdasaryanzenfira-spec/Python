import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from reader import EXPECTED_HEADERS


HEADER_COLOR = "2E7D32"

HEADER_FILL = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

COLUMN_PADDING = 2


def _pretty_header(header):
    return header.replace("_", " ").title()


def autofit_columns(worksheet, headers, students):
    for col_index, header in enumerate(headers):
        max_length = len(_pretty_header(header))

        for student in students:
            cell_length = len(str(student[header]))
            if cell_length > max_length:
                max_length = cell_length

        column_letter = worksheet.cell(row=1, column=col_index + 1).column_letter
        worksheet.column_dimensions[column_letter].width = max_length + COLUMN_PADDING


def write_sheet(worksheet, students):
    pretty_headers = [_pretty_header(h) for h in EXPECTED_HEADERS]
    worksheet.append(pretty_headers)

    for col_index in range(len(EXPECTED_HEADERS)):
        cell = worksheet.cell(row=1, column=col_index + 1)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for student in students:
        worksheet.append([student[header] for header in EXPECTED_HEADERS])

    autofit_columns(worksheet, EXPECTED_HEADERS, students)

    worksheet.sheet_view.showGridLines = True


def create_excel_report(all_students, filtered_students, output_path, is_filtered):
    workbook = Workbook()

    all_sheet = workbook.active
    all_sheet.title = "All Students"
    write_sheet(all_sheet, all_students)

    sheet_titles = ["All Students"]

    if is_filtered:
        filtered_sheet = workbook.create_sheet(title="Filtered")
        write_sheet(filtered_sheet, filtered_students)
        sheet_titles.append("Filtered")

    try:
        workbook.save(output_path)
    except OSError as exc:
        sys.exit(f"Error: could not write output file '{output_path}': {exc}")

    return sheet_titles
